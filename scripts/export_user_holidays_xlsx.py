"""
Dev/test script: list Timetastic users, pick one by ID, fetch /holidays for a date range,
filter by user, export selected fields to XLSX.

Run from project root (folder containing `src/`):

    python scripts/export_user_holidays_xlsx.py

Optional arguments:

    python scripts/export_user_holidays_xlsx.py --start 2024-01-01 --end 2026-12-31
    python scripts/export_user_holidays_xlsx.py --user-id 12345 --output ./my_export.xlsx

Requires TIMETASTIC_API_TOKEN (and optional TIMETASTIC_BASE_URL) in `.env`.

Row colors: red = updatedAt after absence end (retroactive edit); yellow = absence starts
the day after createdAt. See sheet "legenda".
"""

from __future__ import annotations

import argparse
import re
import sys
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

_DATE_ONLY = re.compile(r"^(\d{4})-(\d{2})-(\d{2})$")

# Project root: scripts/ -> parent = repo root
ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.config import load_settings
from src.services.timetastic_service import TimetasticService

# Row highlights (ARGB)
_FILL_RED_RETROACTIVE = PatternFill(fill_type="solid", start_color="FFFFCDCD", end_color="FFFFCDCD")
_FILL_YELLOW_NEXT_DAY = PatternFill(fill_type="solid", start_color="FFFFFFCC", end_color="FFFFFFCC")


def _get(d: Dict[str, Any], *keys: str, default: Any = None) -> Any:
    for k in keys:
        if k in d and d[k] is not None:
            return d[k]
    return default


def pick_date_field(entry: Dict[str, Any], *key_candidates: str) -> str:
    """First non-empty key, formatted as YYYY-MM-DD (API often uses startDate/endDate; OpenAPI sample uses startTime/endTime)."""
    for k in key_candidates:
        if k not in entry:
            continue
        raw = entry[k]
        if raw is None or raw == "":
            continue
        return to_date_only(raw)
    return ""


def format_bool_cell(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, bool):
        return "Yes" if val else "No"
    if val in (0, 1):
        return "Yes" if val == 1 else "No"
    return str(val)


def _format_leave_type(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, str):
        return val
    if isinstance(val, dict):
        return str(
            val.get("name")
            or val.get("Name")
            or val.get("leaveType")
            or val
        )
    return str(val)


def _format_cell(val: Any) -> Any:
    if val is None:
        return ""
    if isinstance(val, (dict, list)):
        return str(val)
    return val


def to_date_only(val: Any) -> str:
    """Return YYYY-MM-DD or empty string; strips time from datetimes and ISO strings."""
    if val is None or val == "":
        return ""
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date) and not isinstance(val, datetime):
        return val.isoformat()
    s = str(val).strip()
    if not s:
        return ""
    if _DATE_ONLY.match(s):
        return s
    try:
        if s.endswith("Z"):
            normalized = s[:-1] + "+00:00"
        else:
            normalized = s
        dt = datetime.fromisoformat(normalized.replace("Z", "+00:00"))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.date().isoformat()
    except ValueError:
        pass
    if "T" in s:
        return s.split("T", 1)[0]
    return s


def normalize_api_range_start(user_input: str) -> str:
    """Date-only '2024-01-01' -> start of UTC day; full ISO left as-is for API."""
    s = user_input.strip()
    m = _DATE_ONLY.match(s)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}T00:00:00Z"
    return s


def normalize_api_range_end(user_input: str) -> str:
    """Date-only '2025-12-31' -> end of UTC day; full ISO left as-is."""
    s = user_input.strip()
    m = _DATE_ONLY.match(s)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}T23:59:59Z"
    return s


def fetch_holidays_raw(service: TimetasticService, start_iso: str, end_iso: str) -> List[Dict[str, Any]]:
    """Paginated GET /holidays (raw JSON objects, no Absence conversion)."""
    raw: List[Dict[str, Any]] = []
    page = 1
    params_base = {"Start": start_iso, "End": end_iso}
    while True:
        params = {**params_base, "PageNumber": page}
        data = service._make_request("/holidays", params=params)
        if isinstance(data, list):
            holidays = data
        else:
            holidays = data.get("holidays", []) or data.get("items", []) or []
        if not holidays:
            break
        raw.extend(holidays)
        if len(holidays) < 100:
            break
        page += 1
    return raw


def filter_by_user_id(rows: List[Dict[str, Any]], user_id: int) -> List[Dict[str, Any]]:
    wanted = int(user_id)
    out: List[Dict[str, Any]] = []
    for entry in rows:
        uid = _get(entry, "userId", "UserId")
        if uid is None and entry.get("user"):
            u = entry["user"]
            if isinstance(u, dict):
                uid = u.get("id") or u.get("Id")
        try:
            if uid is not None and int(uid) == wanted:
                out.append(entry)
        except (TypeError, ValueError):
            continue
    return out


def _parse_iso_date_str(s: str) -> Optional[date]:
    if not s or not _DATE_ONLY.match(s.strip()):
        return None
    try:
        return date.fromisoformat(s.strip())
    except ValueError:
        return None


def row_highlight_fill(entry: Dict[str, Any]) -> Optional[PatternFill]:
    """
    Red: last edit (updatedAt) is after the end of the absence — retroactive change.
    Yellow: absence starts on the calendar day after the record was created.
    Red wins if both match.
    """
    end_s = pick_date_field(
        entry,
        "endDate",
        "EndDate",
        "endTime",
        "EndTime",
        "toDate",
        "ToDate",
    )
    upd_s = pick_date_field(entry, "updatedAt", "UpdatedAt", "updated_at")
    start_s = pick_date_field(
        entry,
        "startDate",
        "StartDate",
        "startTime",
        "StartTime",
        "fromDate",
        "FromDate",
    )
    created_s = pick_date_field(entry, "createdAt", "CreatedAt", "created_at")

    end_d = _parse_iso_date_str(end_s)
    upd_d = _parse_iso_date_str(upd_s)
    start_d = _parse_iso_date_str(start_s)
    created_d = _parse_iso_date_str(created_s)

    if end_d is not None and upd_d is not None and upd_d > end_d:
        return _FILL_RED_RETROACTIVE
    if start_d is not None and created_d is not None and start_d == created_d + timedelta(days=1):
        return _FILL_YELLOW_NEXT_DAY
    return None


def row_to_export(entry: Dict[str, Any]) -> Dict[str, Any]:
    """Map API record to columns; dates as YYYY-MM-DD (start/end from startDate/endDate when API omits startTime/endTime)."""
    return {
        "startTime": pick_date_field(
            entry,
            "startDate",
            "StartDate",
            "startTime",
            "StartTime",
            "fromDate",
            "FromDate",
        ),
        "endTime": pick_date_field(
            entry,
            "endDate",
            "EndDate",
            "endTime",
            "EndTime",
            "toDate",
            "ToDate",
        ),
        "leaveType": _format_leave_type(_get(entry, "leaveType", "LeaveType", "leave_type")),
        "userId": _format_cell(_get(entry, "userId", "UserId")),
        "actionerId": _format_cell(_get(entry, "actionerId", "ActionerId", "actioner_id")),
        "actionedAt": pick_date_field(
            entry,
            "actionedAt",
            "ActionedAt",
            "actioned_at",
        ),
        "createdAt": pick_date_field(
            entry,
            "createdAt",
            "CreatedAt",
            "created_at",
        ),
        "updatedAt": pick_date_field(
            entry,
            "updatedAt",
            "UpdatedAt",
            "updated_at",
        ),
        "leaveTypeAppliesToMaxAbsence": format_bool_cell(
            _get(
                entry,
                "leaveTypeAppliesToMaxAbsence",
                "LeaveTypeAppliesToMaxAbsence",
                "leave_type_applies_to_max_absence",
            )
        ),
        "calendarVisibility": _format_cell(
            _get(entry, "calendarVisibility", "CalendarVisibility", "calendar_visibility")
        ),
        "adjustedEndDate": pick_date_field(
            entry,
            "adjustedEndDate",
            "AdjustedEndDate",
            "adjusted_end_date",
            "endDateWithAdjustment",
            "EndDateWithAdjustment",
        ),
        "isInFuture": format_bool_cell(_get(entry, "isInFuture", "IsInFuture")),
        "isHolidayCancelledOrDeclined": format_bool_cell(
            _get(entry, "isHolidayCancelledOrDeclined", "IsHolidayCancelledOrDeclined")
        ),
    }


def _apply_sheet_layout(ws, columns: List[str]) -> None:
    """Wider columns, bold header, freeze top row."""
    header_font = Font(bold=True, size=11)
    widths = {
        "startTime": 16,
        "endTime": 16,
        "leaveType": 28,
        "userId": 12,
        "actionerId": 12,
        "actionedAt": 16,
        "createdAt": 16,
        "updatedAt": 16,
        "leaveTypeAppliesToMaxAbsence": 28,
        "calendarVisibility": 22,
        "adjustedEndDate": 22,
        "isInFuture": 14,
        "isHolidayCancelledOrDeclined": 28,
    }
    for idx, col_name in enumerate(columns, start=1):
        letter = get_column_letter(idx)
        w = widths.get(col_name, 20)
        ws.column_dimensions[letter].width = max(w, len(col_name) + 2)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def _add_legend_sheet(wb: Workbook) -> None:
    leg = wb.create_sheet("legenda", 1)
    leg["A1"] = "Kolor wiersza (holidays)"
    leg["A2"] = "Czerwony"
    leg["B2"] = "Data ostatniej edycji (updatedAt) jest późniejsza niż koniec nieobecności — zmiana „wstecz” po zakończeniu urlopu."
    leg["A3"] = "Żółty"
    leg["B3"] = "Początek nieobecności przypada na następny dzień kalendarzowy po utworzeniu wpisu (createdAt)."
    leg.column_dimensions["A"].width = 14
    leg.column_dimensions["B"].width = 88
    for coord in ("A1", "A2", "A3"):
        leg[coord].font = Font(bold=True)
    for coord in ("B2", "B3"):
        leg[coord].alignment = Alignment(wrap_text=True, vertical="top")


def print_users(service: TimetasticService) -> None:
    users = service.get_users(force_refresh=False)
    print("\nUsers (id | name | email):\n")
    for u in users:
        uid = u.get("id") or u.get("Id")
        name = u.get("fullName") or u.get("FullName") or u.get("name") or u.get("Name") or ""
        email = u.get("email") or u.get("Email") or ""
        print(f"  {uid}\t{name}\t{email}")
    print(f"\nTotal: {len(users)}\n")


def default_date_range(settings) -> tuple[str, str]:
    if settings.timetastic_test_start_date and settings.timetastic_test_end_date:
        return (
            normalize_api_range_start(settings.timetastic_test_start_date),
            normalize_api_range_end(settings.timetastic_test_end_date),
        )
    now = datetime.now(timezone.utc).year
    start = normalize_api_range_start(f"{now - 2}-01-01")
    end = normalize_api_range_end(f"{now + 1}-12-31")
    return start, end


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Export one user's Timetastic holidays to XLSX.")
    p.add_argument(
        "--start",
        help="Range start: date YYYY-MM-DD or full ISO (date-only is expanded to 00:00:00 UTC)",
    )
    p.add_argument(
        "--end",
        help="Range end: date YYYY-MM-DD or full ISO (date-only is expanded to 23:59:59 UTC)",
    )
    p.add_argument("--user-id", type=int, help="Timetastic user id (skip interactive prompt)")
    p.add_argument(
        "--output",
        "-o",
        help="Output .xlsx path (default: exports/timetastic_user_holidays_<id>_<timestamp>.xlsx)",
    )
    return p.parse_args()


def main() -> None:
    args = parse_args()
    settings = load_settings()
    if not settings.timetastic_api_token:
        print("Error: TIMETASTIC_API_TOKEN is not set. Configure .env in the project root.", file=sys.stderr)
        sys.exit(1)

    service = TimetasticService(settings, storage=None)

    print_users(service)

    user_id: Optional[int] = args.user_id
    if user_id is None:
        raw_in = input("Enter Timetastic user id (number): ").strip()
        try:
            user_id = int(raw_in)
        except ValueError:
            print("Invalid user id.", file=sys.stderr)
            sys.exit(1)

    start_iso, end_iso = default_date_range(settings)
    if args.start:
        start_iso = normalize_api_range_start(args.start)
    if args.end:
        end_iso = normalize_api_range_end(args.end)

    d0 = to_date_only(start_iso)
    d1 = to_date_only(end_iso)
    print(f"Fetching /holidays for dates {d0} .. {d1} (API: {start_iso} .. {end_iso}) ...")
    all_rows = fetch_holidays_raw(service, start_iso, end_iso)
    user_rows = filter_by_user_id(all_rows, user_id)
    print(f"Total API rows in range: {len(all_rows)}, rows for user {user_id}: {len(user_rows)}")

    columns = [
        "startTime",
        "endTime",
        "leaveType",
        "userId",
        "actionerId",
        "actionedAt",
        "createdAt",
        "updatedAt",
        "leaveTypeAppliesToMaxAbsence",
        "calendarVisibility",
        "adjustedEndDate",
        "isInFuture",
        "isHolidayCancelledOrDeclined",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "holidays"
    ws.append(columns)
    for i, entry in enumerate(user_rows):
        r = row_to_export(entry)
        ws.append([r[c] for c in columns])
        row_idx = i + 2
        fill = row_highlight_fill(entry)
        if fill:
            for col_ix in range(1, len(columns) + 1):
                ws.cell(row=row_idx, column=col_ix).fill = fill
    _apply_sheet_layout(ws, columns)
    _add_legend_sheet(wb)

    exports_dir = Path(settings.exports_dir).expanduser()
    exports_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    if args.output:
        out_path = Path(args.output).expanduser()
    else:
        out_path = exports_dir / f"timetastic_user_holidays_{user_id}_{ts}.xlsx"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Wrote: {out_path.resolve()}")


if __name__ == "__main__":
    main()
