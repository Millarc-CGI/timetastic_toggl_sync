"""
File storage for exports and reports.
"""

import json
import os
from typing import List, Dict, Any, Optional
from datetime import datetime, date, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..config import Settings
from ..models.user import User
from ..models.time_entry import TimeEntry
from ..models.absence import Absence
from ..models.report import MonthlyReport, UserReport


class FileStorage:
    """File storage for exports and reports."""
    
    def __init__(self, settings: Settings):
        self.settings = settings
        self.exports_dir = Path(settings.exports_dir)
        self.exports_dir.mkdir(parents=True, exist_ok=True)
    
    def _ensure_month_dir(self, year: int, month: int) -> Path:
        """Ensure month directory exists and return path."""
        month_dir = self.exports_dir / f"{year:04d}-{month:02d}"
        month_dir.mkdir(parents=True, exist_ok=True)
        return month_dir
    
    def _get_role_file_path(self, role: str, year: int, month: int, file_type: str = "csv") -> Path:
        """Get file path for role-based report."""
        month_dir = self._ensure_month_dir(year, month)
        return month_dir / f"{role}_{year:04d}-{month:02d}.{file_type}"
    
    def _get_user_file_path(self, user_email: str, year: int, month: int, file_type: str = "csv") -> Path:
        """Get file path for user-specific report."""
        month_dir = self._ensure_month_dir(year, month)
        # Sanitize email for filename
        safe_email = user_email.replace('@', '_at_').replace('.', '_')
        return month_dir / f"user_{safe_email}_{year:04d}-{month:02d}.{file_type}"
    
    # JSON exports (raw data backup)
    def export_raw_data(
        self, 
        time_entries: List[TimeEntry], 
        absences: List[Absence], 
        year: int, 
        month: int,
        prefix: str = "raw"
    ) -> Path:
        """Export raw data to JSON file."""
        month_dir = self._ensure_month_dir(year, month)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        filename = f"{prefix}_{year:04d}-{month:02d}_{timestamp}.json"
        file_path = month_dir / filename
        
        data = {
            'export_info': {
                'timestamp': datetime.now().isoformat(),
                'year': year,
                'month': month,
                'time_entries_count': len(time_entries),
                'absences_count': len(absences)
            },
            'time_entries': [entry.to_dict() for entry in time_entries],
            'absences': [absence.to_dict() for absence in absences]
        }
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2, default=str)
        
        return file_path
    
    def export_user_report_xlsx(self, report: UserReport) -> Path:
        """Export user report (weekly or monthly) to XLSX with formatting."""
        file_path = self._get_user_file_path(report.user_email, report.year, report.month, "xlsx")
        
        if file_path.exists():
            try:
                os.chmod(file_path, 0o666)
                file_path.unlink()
            except (PermissionError, OSError):
                pass
        
        wb = Workbook()
        self._build_user_report_sheet(wb, report, is_first_sheet=True)
        wb.save(file_path)
        return file_path

    def export_user_reports_xlsx_combined(self, reports: List[UserReport]) -> Path:
        """Export multiple user reports into a single XLSX (one sheet per user)."""
        if not reports:
            raise ValueError("No user reports provided for export.")
        
        year = reports[0].year
        month = reports[0].month
        file_path = self._get_role_file_path("user_combined_database", year, month, "xlsx")
        
        if file_path.exists():
            try:
                os.chmod(file_path, 0o666)
                file_path.unlink()
            except (PermissionError, OSError):
                pass
        
        wb = Workbook()
        first = True
        for report in reports:
            self._build_user_report_sheet(wb, report, is_first_sheet=first)
            first = False
        wb.save(file_path)
        return file_path

    def _build_user_report_sheet(self, wb: Workbook, report: UserReport, is_first_sheet: bool = False):
        """Build a worksheet for a single user report."""
        sheet_title = self._generate_unique_sheet_name(wb, report.user_name or report.user_email or "User")
        ws = wb.active if is_first_sheet else wb.create_sheet()
        ws.title = sheet_title
        
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_font = Font(bold=True, size=14)
        section_font = Font(bold=True, size=11)
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        row = 1
        
        ws.merge_cells(f'A{row}:L{row}')
        title_cell = ws[f'A{row}']
        title_cell.value = f"{report.report_type.title()} User Report"
        title_cell.font = title_font
        title_cell.alignment = center_align
        row += 1
        
        ws[f'A{row}'] = 'User:'
        ws[f'B{row}'] = report.user_name
        row += 1
        ws[f'A{row}'] = 'Email:'
        ws[f'B{row}'] = report.user_email
        row += 1
        ws[f'A{row}'] = 'Department:'
        ws[f'B{row}'] = report.department or 'N/A'
        row += 1
        ws[f'A{row}'] = 'Period:'
        ws[f'B{row}'] = report.period_string
        row += 1
        ws[f'A{row}'] = 'Generated:'
        ws[f'B{row}'] = report.generated_at.strftime("%Y-%m-%d %H:%M:%S") if report.generated_at else ''
        row += 2
        
        ws[f'A{row}'] = 'Totals'
        ws[f'A{row}'].font = section_font
        row += 1
        
        ws[f'A{row}'] = 'Total Hours'
        ws[f'B{row}'] = report.total_hours
        ws[f'B{row}'].number_format = '0.00'
        ws[f'B{row}'].alignment = right_align
        row += 1
        
        ws[f'A{row}'] = 'Weekly Overtime'
        ws[f'B{row}'] = report.weekly_overtime
        ws[f'B{row}'].number_format = '0.00'
        ws[f'B{row}'].alignment = right_align
        row += 1
        
        ws[f'A{row}'] = 'Monthly Overtime'
        ws[f'B{row}'] = report.monthly_overtime
        ws[f'B{row}'].number_format = '0.00'
        ws[f'B{row}'].alignment = right_align
        row += 2
        
        if report.projects_worked:
            ws[f'A{row}'] = 'Projects Worked'
            ws[f'A{row}'].font = section_font
            row += 1
            for project in report.projects_worked:
                ws[f'A{row}'] = project
                row += 1
            row += 1
        
        if report.missing_days:
            ws[f'A{row}'] = 'Missing Days'
            ws[f'A{row}'].font = section_font
            row += 1
            for missing_day in report.missing_days:
                ws[f'A{row}'] = missing_day.isoformat()
                row += 1
            row += 1
        
        breakdown_header_row = None
        if report.daily_breakdown:
            ws[f'A{row}'] = 'Daily Overtime Breakdown'
            ws[f'A{row}'].font = section_font
            row += 1
            
            headers = [
                'Date', 'Type', 'Toggl Hours', 'Absences', 'Total Hours', 'Expected Hours', 'Overtimes',
                'Project', 'Project ID', 'Project Hours', 'Task', 'Task ID', 'Task Hours'
            ]
            
            breakdown_header_row = row
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
            row += 1
            
            for day_entry in report.daily_breakdown:
                date_str = day_entry.get('date')
                if isinstance(date_str, date):
                    date_str = date_str.isoformat()
                elif isinstance(date_str, str):
                    try:
                        date_obj = datetime.fromisoformat(date_str.replace('Z', '+00:00')).date()
                        date_str = date_obj.isoformat()
                    except Exception:
                        pass
                
                entry_type = day_entry.get('type', '')
                toggl_hours = day_entry.get('toggl_hours', 0.0)
                absence_desc = day_entry.get('absence_desc') or ''
                total_hours = day_entry.get('total_hours', 0.0)
                expected_hours = day_entry.get('expected_hours', 0.0)
                ot_hours = day_entry.get('hours', 0.0)
                
                projects = day_entry.get('projects', [])
                
                if projects:
                    project_hours_map: Dict[Any, float] = {}
                    for project in projects:
                        project_id = project.get('project_id')
                        if project_id not in project_hours_map:
                            project_hours_map[project_id] = 0.0
                        project_hours_map[project_id] += project.get('hours', 0.0)
                    
                    for project in projects:
                        project_id = project.get('project_id')
                        project_hours_total = project_hours_map.get(project_id, 0.0)
                        task_hours = project.get('hours', 0.0)
                        
                        data_row = [
                            date_str,
                            entry_type,
                            toggl_hours,
                            absence_desc,
                            total_hours,
                            expected_hours if expected_hours > 0 else None,
                            ot_hours,
                            project.get('project_name', ''),
                            project_id if project_id is not None else '',
                            project_hours_total,
                            project.get('task_name', ''),
                            project.get('task_id') if project.get('task_id') is not None else '',
                            task_hours
                        ]
                        
                        for col_idx, value in enumerate(data_row, start=1):
                            cell = ws.cell(row=row, column=col_idx)
                            cell.value = value
                            cell.border = thin_border
                            
                            if col_idx in [3, 5, 6, 7, 10, 13]:
                                if value is not None:
                                    cell.number_format = '0.00'
                                    cell.alignment = right_align
                                else:
                                    cell.alignment = right_align
                            elif col_idx == 2:
                                cell.alignment = center_align
                            elif col_idx == 4:
                                cell.alignment = left_align
                            else:
                                cell.alignment = left_align
                        
                        row += 1
                else:
                    data_row = [
                        date_str,
                        entry_type,
                        toggl_hours,
                        absence_desc,
                        total_hours,
                        expected_hours if expected_hours > 0 else None,
                        ot_hours,
                        '', '', '', '', '', ''
                    ]
                    
                    for col_idx, value in enumerate(data_row, start=1):
                        cell = ws.cell(row=row, column=col_idx)
                        cell.value = value
                        cell.border = thin_border
                        
                        if col_idx in [3, 5, 6, 7]:
                            if value is not None:
                                cell.number_format = '0.00'
                                cell.alignment = right_align
                            else:
                                cell.alignment = right_align
                        elif col_idx == 2:
                            cell.alignment = center_align
                        elif col_idx == 4:
                            cell.alignment = left_align
                        else:
                            cell.alignment = left_align
                    
                    row += 1
        
        ws.column_dimensions['A'].width = 13.0
        ws.column_dimensions['B'].width = 10.4
        ws.column_dimensions['C'].width = 13.0  # Toggl Hours
        ws.column_dimensions['D'].width = 24.0  # Absences
        ws.column_dimensions['E'].width = 13.0  # Total Hours
        ws.column_dimensions['F'].width = 15.6  # Expected Hours
        ws.column_dimensions['G'].width = 11.7  # Overtimes
        ws.column_dimensions['H'].width = 26.0  # Project
        ws.column_dimensions['I'].width = 13.0  # Project ID
        ws.column_dimensions['J'].width = 15.6  # Project Hours
        ws.column_dimensions['K'].width = 26.0  # Task
        ws.column_dimensions['L'].width = 13.0  # Task ID
        ws.column_dimensions['M'].width = 13.0  # Task Hours
        
        if breakdown_header_row:
            ws.freeze_panes = f'A{breakdown_header_row + 1}'

    @staticmethod
    def _generate_unique_sheet_name(wb: Workbook, name: str) -> str:
        """Ensure sheet name is valid, shortened to 31 chars, and unique."""
        invalid_chars = set(r'[]:*?/\\')
        base = ''.join(ch for ch in name if ch not in invalid_chars).strip() or "Sheet"
        base = base[:31]
        
        sheet_name = base
        counter = 1
        existing = set(wb.sheetnames)
        while sheet_name in existing or not sheet_name:
            suffix = f"_{counter}"
            trimmed_base = base[: max(0, 31 - len(suffix))]
            sheet_name = f"{trimmed_base}{suffix}" if trimmed_base else f"Sheet_{counter}"
            counter += 1
        return sheet_name
    
    def export_admin_report_xlsx(self, reports: List[UserReport], year: int, month: int) -> Path:
        """Export admin report (summary of all users) to XLSX with formatting."""
        file_path = self._get_role_file_path("admin", year, month, "xlsx")
        
        # Remove existing file if it exists to allow overwriting
        if file_path.exists():
            try:
                os.chmod(file_path, 0o666)
                file_path.unlink()
            except (PermissionError, OSError):
                pass
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Admin Report"
        
        # Define styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_font = Font(bold=True, size=14)
        summary_font = Font(bold=True, size=11)
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        row = 1
        
        # Title
        ws.merge_cells(f'A{row}:H{row}')
        title_cell = ws[f'A{row}']
        title_cell.value = 'Admin Report - All Users'
        title_cell.font = title_font
        title_cell.alignment = center_align
        row += 1
        
        # Period and metadata
        ws[f'A{row}'] = 'Period:'
        ws[f'B{row}'] = f"{year}-{month:02d}"
        row += 1
        ws[f'A{row}'] = 'Generated:'
        ws[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row += 1
        ws[f'A{row}'] = 'Total Users:'
        ws[f'B{row}'] = len(reports)
        row += 2
        
        # Summary statistics - traktuj ujemne overtime jako 0
        total_hours = sum(r.total_hours for r in reports)
        total_overtime = sum(max(0.0, r.monthly_overtime) for r in reports)
        total_weekend_overtime = sum(max(0.0, r.weekend_overtime) for r in reports)
        expected_hours_per_user = reports[0].expected_hours if reports else 0.0
        
        ws[f'A{row}'] = 'Summary Statistics'
        ws[f'A{row}'].font = summary_font
        row += 1
        
        ws[f'A{row}'] = 'Total Hours (All Users)'
        ws[f'B{row}'] = total_hours
        ws[f'B{row}'].number_format = '0.00'
        ws[f'B{row}'].alignment = right_align
        row += 1

        ws[f'A{row}'] = 'Expected Hours (per user)'
        ws[f'B{row}'] = expected_hours_per_user
        ws[f'B{row}'].number_format = '0.00'
        ws[f'B{row}'].alignment = right_align
        row += 1
        
        ws[f'A{row}'] = 'Total Overtime (All Users)'
        ws[f'B{row}'] = total_overtime
        ws[f'B{row}'].number_format = '0.00'
        ws[f'B{row}'].alignment = right_align
        row += 1
        
        ws[f'A{row}'] = 'Total Weekend Overtime (All Users)'
        ws[f'B{row}'] = total_weekend_overtime
        ws[f'B{row}'].number_format = '0.00'
        ws[f'B{row}'].alignment = right_align
        row += 2
        
        # User Details Table Header
        headers = [
            'User', 'Department', 'Period Label',
            'Total Hours', 'Monthly Overtime', 'Weekend Overtime', 'Working Days', 'Missing Toggl Entries'
        ]
        
        header_row = row
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        row += 1
        
        # User Details Data
        sorted_reports = sorted(reports, key=lambda x: x.total_hours, reverse=True)
        for report in sorted_reports:
            # Traktuj ujemne overtime jako 0
            monthly_ot = max(0.0, report.monthly_overtime)
            weekend_ot = max(0.0, report.weekend_overtime)
            working_days = sum(
                1 for day in (report.daily_breakdown or [])
                if (day.get('toggl_hours') or 0.0) > 0
            )
            
            data_row = [
                report.user_name,
                report.department or 'N/A',
                report.period_label,
                report.total_hours,
                monthly_ot,
                weekend_ot,
                working_days,
                len(report.missing_days)
            ]
            
            for col_idx, value in enumerate(data_row, start=1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = value
                cell.border = thin_border
                
                # Format numbers
                if col_idx in [4, 5, 6]:  # Total Hours, Monthly Overtime, Weekend Overtime
                    cell.number_format = '0.00'
                    cell.alignment = right_align
                elif col_idx == 7:  # Working Days
                    cell.alignment = center_align
                elif col_idx == 8:  # Missing Days
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align
            
            row += 1
        
        # Set column widths for better readability (increased by 30%)
        ws.column_dimensions['A'].width = 32.5  # User (25 * 1.3)
        ws.column_dimensions['B'].width = 26.0  # Department (20 * 1.3)
        ws.column_dimensions['C'].width = 15.6  # Period Label (12 * 1.3)
        ws.column_dimensions['D'].width = 15.6  # Total Hours (12 * 1.3)
        ws.column_dimensions['E'].width = 20.8  # Monthly Overtime (16 * 1.3)
        ws.column_dimensions['F'].width = 23.4  # Weekend Overtime (18 * 1.3)
        ws.column_dimensions['G'].width = 16.0  # Working Days
        ws.column_dimensions['H'].width = 19.5  # Missing Toggl Entries (15 * 1.3, increased for better readability)
        
        # Freeze header row
        ws.freeze_panes = f'A{header_row + 1}'
        
        # Per-user sheets with simplified daily breakdown
        for report in sorted_reports:
            sheet_name = self._generate_unique_sheet_name(wb, report.user_name or report.user_email or "User")
            user_ws = wb.create_sheet(title=sheet_name)
            
            user_headers = [
                'Date', 'Type', 'Toggl Hours', 'Absences', 'Total Hours',
                'Expected Hours', 'Overtime', 'Weekend Overtime', 'Missing Toggl Entries'
            ]
            
            # Header row
            for col_idx, header in enumerate(user_headers, start=1):
                cell = user_ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
            
            missing_set = {d.isoformat() for d in (report.missing_days or [])}
            row_idx = 2
            for day in report.daily_breakdown or []:
                date_str = day.get('date')
                if isinstance(date_str, date):
                    date_str = date_str.isoformat()
                elif isinstance(date_str, str):
                    try:
                        date_obj = datetime.fromisoformat(date_str.replace('Z', '+00:00')).date()
                        date_str = date_obj.isoformat()
                    except Exception:
                        pass
                
                entry_type = day.get('type', '')
                type_label = 'weekend' if entry_type == 'weekend' else ''
                toggl_hours = day.get('toggl_hours', 0.0)
                absences = day.get('absence_desc') or ''
                total_hours = day.get('total_hours', 0.0)
                expected_hours = day.get('expected_hours', 0.0) if entry_type != 'weekend' else None
                ot_hours = day.get('hours', 0.0)
                overtime = ot_hours if entry_type != 'weekend' else None
                weekend_overtime = ot_hours if entry_type == 'weekend' else None
                missing_flag = 1 if date_str in missing_set else ''
                
                data_row = [
                    date_str,
                    type_label,
                    toggl_hours,
                    absences,
                    total_hours,
                    expected_hours,
                    overtime,
                    weekend_overtime,
                    missing_flag
                ]
                
                for col_idx, value in enumerate(data_row, start=1):
                    cell = user_ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.border = thin_border
                    
                    if col_idx in [3, 5, 6, 7, 8]:  # numeric columns
                        if value not in (None, ''):
                            cell.number_format = '0.00'
                            cell.alignment = right_align
                        else:
                            cell.alignment = right_align
                    elif col_idx == 2:
                        cell.alignment = center_align
                    else:
                        cell.alignment = left_align
                
                row_idx += 1
            
            # Column widths
            user_ws.column_dimensions['A'].width = 13.0
            user_ws.column_dimensions['B'].width = 10.0
            user_ws.column_dimensions['C'].width = 13.0
            user_ws.column_dimensions['D'].width = 24.0
            user_ws.column_dimensions['E'].width = 13.0
            user_ws.column_dimensions['F'].width = 15.0
            user_ws.column_dimensions['G'].width = 12.0
            user_ws.column_dimensions['H'].width = 16.0
            user_ws.column_dimensions['I'].width = 16.0
            
            user_ws.freeze_panes = "A2"
        
        # Save workbook
        wb.save(file_path)
        
        return file_path
    
    def export_monthly_project_stats_xlsx(
        self,
        records: List[Dict[str, Any]],
        year: int,
        month: int
    ) -> Path:
        """Export monthly project statistics to XLSX in monthly folder."""
        file_path = self._get_role_file_path("project_stats", year, month, "xlsx")
        
        # Remove existing file if it exists to allow overwriting
        if file_path.exists():
            try:
                os.chmod(file_path, 0o666)
                file_path.unlink()
            except (PermissionError, OSError):
                pass
        
        headers = [
            "User",
            "Project",
            "Tasks",
            "Total Hours",
            "Normal Overtime",
            "Weekend Overtime",
        ]
        
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Project Stats"
        
        # Define styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_font = Font(bold=True, size=14)
        center_align = Alignment(horizontal="center", vertical="center")
        
        row = 1
        
        # Title
        sheet.merge_cells(f'A{row}:F{row}')
        title_cell = sheet[f'A{row}']
        title_cell.value = f'Monthly Project Statistics - {year}-{month:02d}'
        title_cell.font = title_font
        title_cell.alignment = center_align
        row += 2
        
        # Add headers
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
        row += 1
        
        # Add data rows
        for record in records:
            sheet.append([
                record.get("user_name") or record.get("user_email"),
                record.get("project"),
                record.get("tasks_text") or record.get("task"),
                round(record.get("total_hours", 0.0), 2),
                round(record.get("normal_overtime", 0.0), 2),
                round(record.get("weekend_overtime", 0.0), 2),
            ])
        
        # Add summary row
        if records:
            summary_row = sheet.max_row + 1
            total_hours = round(sum(r.get("total_hours", 0.0) for r in records), 2)
            total_normal_ot = round(sum(r.get("normal_overtime", 0.0) for r in records), 2)
            total_weekend_ot = round(sum(r.get("weekend_overtime", 0.0) for r in records), 2)

            sheet.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
            sheet.cell(row=summary_row, column=2, value="").font = Font(bold=True)
            sheet.cell(row=summary_row, column=4, value=total_hours).font = Font(bold=True)
            sheet.cell(row=summary_row, column=5, value=total_normal_ot).font = Font(bold=True)
            sheet.cell(row=summary_row, column=6, value=total_weekend_ot).font = Font(bold=True)
        
        # Auto-size columns (30% wider)
        for column_idx in range(1, len(headers) + 1):
            column_letter = get_column_letter(column_idx)
            max_length = 0
            for cell in sheet[column_letter]:
                if cell.value is None:
                    continue
                max_length = max(max_length, len(str(cell.value)))
            base_width = min(max(max_length + 2, 12), 60)
            sheet.column_dimensions[column_letter].width = base_width * 1.3
        
        workbook.save(file_path)
        return file_path
    
    # Utility methods
    def list_available_reports(self, year: int, month: int) -> Dict[str, List[Path]]:
        """List available report files for a given month."""
        month_dir = self._ensure_month_dir(year, month)
        
        reports = {
            'admin': [],
            'user': [],
            'raw': []
        }
        
        for file_path in month_dir.glob("*.csv"):
            filename = file_path.name
            
            if filename.startswith("admin_"):
                reports['admin'].append(file_path)
            elif filename.startswith("user_"):
                reports['user'].append(file_path)
        
        for file_path in month_dir.glob("*.json"):
            filename = file_path.name
            if filename.startswith("raw_"):
                reports['raw'].append(file_path)
        
        return reports
    
    def get_report_file_info(self, file_path: Path) -> Dict[str, Any]:
        """Get information about a report file."""
        try:
            stat = file_path.stat()
            return {
                'path': str(file_path),
                'size_bytes': stat.st_size,
                'size_mb': stat.st_size / (1024 * 1024),
                'created': datetime.fromtimestamp(stat.st_ctime).isoformat(),
                'modified': datetime.fromtimestamp(stat.st_mtime).isoformat(),
                'exists': True
            }
        except FileNotFoundError:
            return {
                'path': str(file_path),
                'exists': False
            }
    
    def cleanup_old_reports(self, months_to_keep: int = 12):
        """Clean up old report files."""
        cutoff_date = datetime.now() - timedelta(days=months_to_keep * 30)
        
        deleted_count = 0
        for month_dir in self.exports_dir.iterdir():
            if not month_dir.is_dir():
                continue
            
            try:
                # Extract year-month from directory name
                year_month = month_dir.name
                if '-' in year_month and len(year_month.split('-')) == 2:
                    year_str, month_str = year_month.split('-')
                    year = int(year_str)
                    month = int(month_str)
                    
                    # Check if this month is older than cutoff
                    month_date = datetime(year, month, 1)
                    if month_date < cutoff_date:
                        # Delete the entire month directory
                        import shutil
                        shutil.rmtree(month_dir)
                        deleted_count += 1
                        print(f"Deleted old reports: {month_dir}")
            except (ValueError, OSError) as e:
                print(f"Error processing directory {month_dir}: {e}")
        
        print(f"Cleanup completed: {deleted_count} month directories deleted")
        return deleted_count
