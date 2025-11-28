"""
Statistics generator for creating various statistical reports.
"""

from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime, date, timedelta
from collections import defaultdict, Counter

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from ..config import Settings
from ..models.user import User
from ..models.project import Project


class StatisticsGenerator:
    """Generates various statistics and reports."""
    
    def __init__(self, settings: Settings):
        self.settings = settings

    @staticmethod
    def normalize_project_key(name: Optional[str]) -> str:
        return (name or "").strip().lower()

    def _resolve_project_token(
        self,
        token: str,
        by_id: Dict[int, Project],
        by_name: Dict[str, Project],
    ) -> Optional[Project]:
        token = (token or "").strip()
        if not token:
            return None
        project = None
        try:
            pid = int(token)
            project = by_id.get(pid)
        except ValueError:
            project = by_name.get(self.normalize_project_key(token))
        return project

    def build_project_activity_map(
        self,
        projects: List[Project],
        time_entries: List[Any],
        lookback_start: date,
        lookback_end: date,
    ) -> Dict[int, Dict[str, Any]]:
        """Return per-project stats (activity window + hours) within lookback."""
        id_map = {p.project_id: p for p in projects if p.project_id is not None}
        name_map = {
            self.normalize_project_key(p.name): p
            for p in projects
            if p.name and p.project_id is not None
        }
        activity: Dict[int, Dict[str, Any]] = {}
        for entry in time_entries:
            project = id_map.get(entry.project_id)
            if not project and entry.project_name:
                project = name_map.get(self.normalize_project_key(entry.project_name))
            if not project or not project.active:
                continue
            project_id = project.project_id
            stat = activity.setdefault(
                project_id,
                {
                    'project': project,
                    'project_id': project_id,
                    'start': None,
                    'end': None,
                    'total_hours': 0.0,
                    'normalized_name': self.normalize_project_key(project.name or f"Project {project_id}"),
                },
            )
            day = entry.date
            if stat['start'] is None or day < stat['start']:
                stat['start'] = day
            if stat['end'] is None or day > stat['end']:
                stat['end'] = day
            stat['total_hours'] += entry.duration_hours
        # Ensure start/end at least within lookback
        for stat in activity.values():
            if stat['start'] is None:
                stat['start'] = lookback_start
            if stat['end'] is None:
                stat['end'] = lookback_end
        return activity

    @staticmethod
    def build_project_windows(
        selected_activity: List[Dict[str, Any]],
        lookback_start: date,
        lookback_end: date,
    ) -> Dict[str, Dict[Any, Dict[str, date]]]:
        """Return lookup tables for project windows."""
        by_id: Dict[int, Dict[str, date]] = {}
        by_name: Dict[str, Dict[str, date]] = {}
        for item in selected_activity:
            window_start = item.get('start') or lookback_start
            window_end = item.get('end') or lookback_end
            if window_end < window_start:
                window_end = window_start
            project_id = item['project_id']
            normalized = item.get('normalized_name') or ""
            window = {'start': window_start, 'end': window_end}
            by_id[project_id] = window
            if normalized:
                by_name[normalized] = window
        return {'by_id': by_id, 'by_name': by_name}

    def filter_time_entries_for_projects(
        self,
        time_entries: List[Any],
        windows: Dict[str, Dict[Any, Dict[str, date]]],
    ) -> Tuple[List[Any], Counter]:
        """Filter entries so they fall inside selected project windows."""
        filtered: List[Any] = []
        stats: Counter = Counter()
        by_id = windows.get('by_id', {})
        by_name = windows.get('by_name', {})
        for entry in time_entries:
            window = None
            if entry.project_id is not None and entry.project_id in by_id:
                window = by_id[entry.project_id]
            elif entry.project_name:
                normalized = self.normalize_project_key(entry.project_name)
                window = by_name.get(normalized)
            if not window:
                stats['no_match'] += 1
                continue
            day = entry.date
            if window['start'] <= day <= window['end']:
                filtered.append(entry)
                stats['match'] += 1
            else:
                stats['out_of_window'] += 1
        return filtered, stats

    def compute_project_precise_totals(
        self,
        entries_by_email: Dict[str, List[Any]],
        project_info: Dict[str, Any],
        default_start: date,
        default_end: date,
    ) -> Dict[str, float]:
        """Roll raw Toggl durations per user for a project window."""
        project_id = project_info.get('project_id')
        normalized = project_info.get('normalized_name') or ""
        window_start = project_info.get('start') or default_start
        window_end = project_info.get('end') or default_end
        precise: Dict[str, float] = {}
        for email, entries in entries_by_email.items():
            total = 0.0
            for entry in entries:
                day = entry.date
                if day < window_start or day > window_end:
                    continue
                matches = False
                if project_id is not None and entry.project_id == project_id:
                    matches = True
                elif normalized and self.normalize_project_key(entry.project_name) == normalized:
                    matches = True
                if matches:
                    total += entry.duration_hours
            if total > 0:
                precise[email] = total
        return precise
    
    def generate_user_stats(
        self,
        user_email: str,
        user_data: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Generate statistics for a specific user."""
        
        daily_data = user_data.get('daily_data', [])
        if not daily_data:
            return {
                'user_email': user_email,
                'error': 'No daily data available'
            }
        
        # Basic statistics
        total_hours = user_data.get('total_hours', 0)
        billable_hours = user_data.get('billable_hours', 0)
        working_days = user_data.get('working_days', 0)
        missing_days = len(user_data.get('missing_days', []))
        
        # Daily statistics
        daily_hours = [day['total_hours'] for day in daily_data]
        billable_daily = [day['billable_hours'] for day in daily_data]
        
        # Project statistics
        project_hours = user_data.get('project_hours', {})
        projects_worked = len(project_hours)
        top_project = max(project_hours.items(), key=lambda x: x[1]) if project_hours else None
        
        # Absence statistics
        absence_breakdown = user_data.get('absence_breakdown', {})
        total_absence_days = sum(absence_breakdown.values())
        
        # Time distribution
        hours_by_day_of_week = defaultdict(float)
        for day_data in daily_data:
            day_of_week = day_data['date'].weekday()  # Monday = 0, Sunday = 6
            hours_by_day_of_week[day_of_week] += day_data['total_hours']
        
        # Calculate averages
        avg_daily_hours = sum(daily_hours) / len(daily_hours) if daily_hours else 0
        avg_billable_hours = sum(billable_daily) / len(billable_daily) if billable_daily else 0
        
        # Productivity metrics
        productivity_score = 0
        if total_hours > 0:
            # Simple productivity score based on billable ratio and consistency
            billable_ratio = billable_hours / total_hours
            consistency = 1 - (max(daily_hours) - min(daily_hours)) / max(daily_hours) if max(daily_hours) > 0 else 1
            productivity_score = (billable_ratio * 0.7 + consistency * 0.3) * 100
        
        return {
            'user_email': user_email,
            'period': user_data.get('period_string', 'Unknown'),
            'basic_stats': {
                'total_hours': total_hours,
                'billable_hours': billable_hours,
                'working_days': working_days,
                'missing_days': missing_days,
                'total_absence_days': total_absence_days,
                'projects_worked': projects_worked
            },
            'averages': {
                'avg_daily_hours': avg_daily_hours,
                'avg_billable_hours': avg_billable_hours,
                'avg_hours_per_working_day': total_hours / working_days if working_days > 0 else 0
            },
            'project_stats': {
                'total_projects': projects_worked,
                'top_project': top_project,
                'project_distribution': dict(sorted(project_hours.items(), key=lambda x: x[1], reverse=True))
            },
            'absence_stats': {
                'total_absence_days': total_absence_days,
                'absence_breakdown': absence_breakdown,
                'absence_rate': total_absence_days / len(daily_data) if daily_data else 0
            },
            'time_distribution': {
                'hours_by_day_of_week': dict(hours_by_day_of_week),
                'weekday_avg': sum(hours_by_day_of_week[i] for i in range(5)) / 5 if any(hours_by_day_of_week[i] for i in range(5)) else 0,
                'weekend_avg': sum(hours_by_day_of_week[i] for i in [5, 6]) / 2 if any(hours_by_day_of_week[i] for i in [5, 6]) else 0
            },
            'productivity_metrics': {
                'billable_ratio': billable_hours / total_hours if total_hours > 0 else 0,
                'consistency_score': 1 - (max(daily_hours) - min(daily_hours)) / max(daily_hours) if max(daily_hours) > 0 else 1,
                'productivity_score': productivity_score
            },
            'daily_breakdown': daily_data
        }
    
    def generate_project_stats(
        self,
        all_user_data: Dict[str, Dict[str, Any]],
        users: List[User],
        projects: Optional[List[Project]] = None
    ) -> Dict[Any, Dict[str, Any]]:
        """Generate statistics for all projects, prioritizing Toggl project metadata."""
        
        user_lookup = {user.email.lower(): user for user in users}
        project_lookup = {
            project.project_id: project
            for project in (projects or [])
            if project.project_id is not None and project.active and (project.billable is None or project.billable)
        }
        
        stats_by_id: Dict[Any, Dict[str, Any]] = {}
        for user_email, user_data in all_user_data.items():
            user = user_lookup.get(user_email)
            department = user.department if user else 'Unknown'
            project_hours_by_id = user_data.get('project_hours_by_id') or {}
            project_task_hours_by_id = user_data.get('project_task_hours_by_id') or {}
            activity_dates = user_data.get('project_activity_dates') or {}
            project_names_by_id = user_data.get('project_names_by_id') or {}
            
            for project_id, hours in project_hours_by_id.items():
                if project_id is None:
                    continue
                if project_lookup and project_id not in project_lookup:
                    continue
                project_obj = project_lookup.get(project_id)
                stats = stats_by_id.setdefault(project_id, {
                    'project_id': project_id,
                    'project': project_obj,
                    'project_name': project_obj.name if project_obj else project_names_by_id.get(project_id, f"Project {project_id}"),
                    'total_hours': 0.0,
                    'users': set(),
                    'user_hours': defaultdict(float),
                    'task_hours': defaultdict(float),
                    'departments': set(),
                    'activity_start': None,
                    'activity_end': None,
                })
                stats['total_hours'] += hours
                stats['users'].add(user_email)
                stats['user_hours'][user_email] += hours
                stats['departments'].add(department)
                
                for task, task_hours in (project_task_hours_by_id.get(project_id, {}) or {}).items():
                    stats['task_hours'][task] += task_hours
                
                window = activity_dates.get(project_id) or {}
                start_date = window.get('start')
                end_date = window.get('end')
                if start_date and (stats['activity_start'] is None or start_date < stats['activity_start']):
                    stats['activity_start'] = start_date
                if end_date and (stats['activity_end'] is None or end_date > stats['activity_end']):
                    stats['activity_end'] = end_date

        for project_id, project in project_lookup.items():
            stats_by_id.setdefault(project_id, {
                'project_id': project_id,
                'project': project,
                'project_name': project.name,
                'total_hours': 0.0,
                'users': set(),
                'user_hours': defaultdict(float),
                'task_hours': defaultdict(float),
                'departments': set(),
                'activity_start': project.start_date,
                'activity_end': project.end_date,
            })
        
        # Fallback to name-based aggregation when IDs are unavailable
        if not stats_by_id:
            project_stats = defaultdict(lambda: {
                'total_hours': 0,
                'users': set(),
                'user_hours': defaultdict(float),
                'departments': set(),
                'task_hours': defaultdict(float)
            })
            
            for user_email, user_data in all_user_data.items():
                user = user_lookup.get(user_email)
                department = user.department if user else 'Unknown'
                project_hours = user_data.get('project_hours', {})
                project_task_hours = user_data.get('project_task_hours', {})
                
                for project_name, hours in project_hours.items():
                    project_stats[project_name]['total_hours'] += hours
                    project_stats[project_name]['users'].add(user_email)
                    project_stats[project_name]['user_hours'][user_email] += hours
                    project_stats[project_name]['departments'].add(department)
                    
                    for task, task_hours in (project_task_hours.get(project_name, {}) or {}).items():
                        project_stats[project_name]['task_hours'][task] += task_hours
            
            return {
                project_name: {
                    'project_name': project_name,
                    'project_id': None,
                    'total_hours': stats['total_hours'],
                    'total_users': len(stats['users']),
                    'average_hours_per_user': stats['total_hours'] / len(stats['users']) if stats['users'] else 0,
                    'departments': list(stats['departments']),
                    'user_distribution': dict(stats['user_hours']),
                    'task_distribution': dict(
                        sorted(stats['task_hours'].items(), key=lambda x: x[1], reverse=True)
                    ),
                    'project_start': None,
                    'project_end': None,
                }
                for project_name, stats in project_stats.items()
            }
        
        final_stats: Dict[Any, Dict[str, Any]] = {}
        for project_id, stats in stats_by_id.items():
            total_users = len(stats['users'])
            project_obj = stats.get('project')
            project_start = project_obj.start_date if project_obj and project_obj.start_date else stats.get('activity_start')
            project_end = project_obj.end_date if project_obj and project_obj.end_date else stats.get('activity_end')
            
            final_stats[project_id] = {
                'project_id': project_id,
                'project_name': stats.get('project_name', f"Project {project_id}"),
                'total_hours': stats['total_hours'],
                'total_users': total_users,
                'average_hours_per_user': stats['total_hours'] / total_users if total_users else 0,
                'departments': list(stats['departments']),
                'user_distribution': dict(stats['user_hours']),
                'task_distribution': dict(
                    sorted(stats['task_hours'].items(), key=lambda x: x[1], reverse=True)
                ),
                'project_start': project_start,
                'project_end': project_end,
            }
        
        return final_stats

    def summarize_project_activity(
        self,
        projects: List[Project],
        months: int = 6,
        reference_date: Optional[date] = None
    ) -> List[Dict[str, Any]]:
        """Summarize project activity for the last N months using Toggl metadata."""
        if not projects:
            return []

        today = reference_date or date.today()
        months = max(1, months)
        window_start = today - timedelta(days=months * 30)

        records: List[Dict[str, Any]] = []
        for project in projects:
            start_date = project.start_date or (project.created_at.date() if project.created_at else None)
            end_date = project.end_date
            last_updated = project.updated_at.date() if project.updated_at else None
            effective_end = end_date or last_updated or today
            effective_start = start_date or (project.created_at.date() if project.created_at else effective_end)

            if effective_end < window_start:
                continue

            duration_end = end_date or today
            duration_days = None
            if effective_start:
                duration_days = (duration_end - effective_start).days + 1

            status_label = "Active" if project.active else "Closed"

            records.append({
                'project_id': project.project_id,
                'project_name': project.name,
                'status': status_label,
                'start_date': effective_start,
                'end_date': end_date,
                'duration_days': duration_days,
                'last_updated': last_updated,
                'active_within_window': True,
            })

        def sort_key(record: Dict[str, Any]) -> Any:
            start_ord = (record.get('start_date') or date.min).toordinal()
            return (0 if record.get('status') == 'Active' else 1, -start_ord, record.get('project_name', '').lower())

        records.sort(key=sort_key)
        return records

    def select_projects_by_tokens(
        self,
        projects: List[Project],
        activity_map: Dict[int, Dict[str, Any]],
        requested_tokens: Optional[List[str]],
        default_limit: int,
    ) -> Dict[str, Any]:
        """Resolve requested projects (by ID or name) and report inactive/missing ones."""
        project_by_id = {p.project_id: p for p in projects if p.project_id is not None}
        project_by_name = {
            self.normalize_project_key(p.name): p
            for p in projects
            if p.name and p.project_id is not None
        }
        activity_list = sorted(
            activity_map.values(),
            key=lambda item: item.get('start') or date.min,
            reverse=True,
        )

        selected_activity: List[Dict[str, Any]] = []
        inactive_projects: List[Project] = []
        missing_tokens: List[str] = []
        no_entries_projects: List[Project] = []
        seen_ids: set[int] = set()

        tokens = requested_tokens or []
        if tokens:
            for token in tokens:
                project = self._resolve_project_token(token, project_by_id, project_by_name)
                if not project:
                    missing_tokens.append(token)
                    continue
                if not project.active:
                    inactive_projects.append(project)
                    continue
                if project.project_id in seen_ids:
                    continue
                activity = activity_map.get(project.project_id)
                if not activity:
                    no_entries_projects.append(project)
                    continue
                selected_activity.append(activity)
                seen_ids.add(project.project_id)
        else:
            for activity in activity_list[: default_limit or 3]:
                selected_activity.append(activity)

        return {
            'selected_activity': selected_activity,
            'activity_list': activity_list,
            'inactive_projects': inactive_projects,
            'missing_tokens': missing_tokens,
            'no_entries_projects': no_entries_projects,
        }

    def generate_user_project_task_stats(
        self,
        all_user_data: Dict[str, Dict[str, Any]],
        users: List[User],
        overtime_data: Dict[str, Dict[str, Any]]
    ) -> List[Dict[str, Any]]:
        """Build per-user/project/task statistics with overtime allocation."""
        user_lookup = {user.email.lower(): user for user in users}
        rows: List[Dict[str, Any]] = []

        for user_email, user_snapshot in all_user_data.items():
            user = user_lookup.get(user_email)
            user_name = user.display_name if user else user_email
            user_rows = self._build_user_project_task_rows(
                user_email,
                user_name,
                user_snapshot,
                overtime_data.get(user_email, {}),
                project_filter=None,
                start_date=None,
                end_date=None,
            )
            rows.extend(user_rows)

        return rows

    def generate_project_specific_stats(
        self,
        all_user_data: Dict[str, Dict[str, Any]],
        users: List[User],
        overtime_data: Dict[str, Dict[str, Any]],
        project_name: str,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> List[Dict[str, Any]]:
        """Generate statistics for a specific project window."""
        user_lookup = {user.email.lower(): user for user in users}
        rows: List[Dict[str, Any]] = []

        normalized_project = self._canonical_project_key(project_name)

        for user_email, user_snapshot in all_user_data.items():
            user = user_lookup.get(user_email)
            user_name = user.display_name if user else user_email
            user_rows = self._build_user_project_task_rows(
                user_email,
                user_name,
                user_snapshot,
                overtime_data.get(user_email, {}),
                project_filter=normalized_project,
                start_date=start_date,
                end_date=end_date,
            )
            rows.extend(user_rows)

        return rows

    def export_user_project_task_xlsx(
        self,
        records: List[Dict[str, Any]],
        output_path: str | Path
    ) -> str:
        """Export user/project/task statistics to XLSX and return the path."""
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        headers = [
            "User",
            "Project",
            "Tasks",
            "Total Hours",
            "Normal Overtime",
            "Weekend Overtime",
        ]

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "User Project Stats"

        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for record in records:
            sheet.append([
                record.get("user_name") or record.get("user_email"),
                record.get("project"),
                record.get("tasks_text") or record.get("task"),
                round(record.get("total_hours", 0.0), 2),
                round(record.get("normal_overtime", 0.0), 2),
                round(record.get("weekend_overtime", 0.0), 2),
            ])

        for column_idx in range(1, len(headers) + 1):
            column_letter = get_column_letter(column_idx)
            max_length = 0
            for cell in sheet[column_letter]:
                if cell.value is None:
                    continue
                max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)

        workbook.save(path)
        return str(path)

    def _build_user_project_task_rows(
        self,
        user_email: str,
        user_name: str,
        user_data: Dict[str, Any],
        overtime_data: Dict[str, Any],
        project_filter: Optional[str],
        start_date: Optional[date],
        end_date: Optional[date],
    ) -> List[Dict[str, Any]]:
        """Convert daily data + overtime info into per-project/task rows."""
        store: Dict[str, Dict[str, Any]] = defaultdict(
            lambda: {
                'total_hours': 0.0,
                'normal_overtime': 0.0,
                'weekend_overtime': 0.0,
                'task_hours': defaultdict(float),
            }
        )
        daily_data = user_data.get('daily_data') or []
        normalized_filter = project_filter

        overtime_lookup: Dict[date, Dict[str, float]] = defaultdict(lambda: {'weekday': 0.0, 'weekend': 0.0})
        for item in (overtime_data.get('daily_breakdown') or []):
            day = item.get('date')
            if not day:
                continue
            if item.get('type') == 'weekday':
                overtime_lookup[day]['weekday'] += item.get('hours', 0.0)
            else:
                overtime_lookup[day]['weekend'] += item.get('hours', 0.0)

        for day in daily_data:
            day_date = day.get('date')
            if start_date and day_date and day_date < start_date:
                continue
            if end_date and day_date and day_date > end_date:
                continue

            project_hours = day.get('project_hours') or {}
            if not project_hours:
                continue
            total_project_hours = sum(project_hours.values())
            if total_project_hours <= 0:
                continue

            project_task_hours = day.get('task_hours') or {}
            day_ot = overtime_lookup.get(day.get('date'), {})
            normal_total = day_ot.get('weekday', 0.0)
            weekend_total = day_ot.get('weekend', 0.0)

            for project_name, hours in project_hours.items():
                if hours <= 0:
                    continue
                project_key = self._canonical_project_key(project_name)
                if normalized_filter and project_key != normalized_filter:
                    continue
                display_project = self._display_project_name(project_name)
                project_entry = store[display_project]
                normal_share = normal_total * (hours / total_project_hours) if total_project_hours > 0 else 0.0
                weekend_share = weekend_total * (hours / total_project_hours) if total_project_hours > 0 else 0.0
                project_entry['total_hours'] += hours
                project_entry['normal_overtime'] += normal_share
                project_entry['weekend_overtime'] += weekend_share

                tasks_for_project = (
                    project_task_hours.get(project_name)
                    or project_task_hours.get(display_project)
                    or {}
                )
                task_total = sum(tasks_for_project.values()) if tasks_for_project else 0.0
                task_bucket = project_entry['task_hours']

                if tasks_for_project:
                    for task_name, task_hours in tasks_for_project.items():
                        if task_hours <= 0:
                            continue
                        task_bucket[task_name or "No Task"] += task_hours
                    remainder = max(0.0, hours - task_total)
                    if remainder > 1e-3:
                        task_bucket["No Task"] += remainder
                else:
                    task_bucket["No Task"] += hours

        results: List[Dict[str, Any]] = []
        for project, metrics in store.items():
            if (
                metrics['total_hours'] <= 0
                and metrics['normal_overtime'] <= 0
                and metrics['weekend_overtime'] <= 0
            ):
                continue
            task_hours_map = metrics.get('task_hours', {})
            task_names = sorted(task_hours_map.keys(), key=lambda key: task_hours_map[key], reverse=True)
            tasks_text = ", ".join(task_names) if task_names else "No Task"
            results.append({
                'user_email': user_email,
                'user_name': user_name,
                'project': project,
                'task': tasks_text,
                'tasks': task_names,
                'tasks_text': tasks_text,
                'total_hours': metrics['total_hours'],
                'normal_overtime': metrics['normal_overtime'],
                'weekend_overtime': metrics['weekend_overtime'],
            })
        return results

    @staticmethod
    def _canonical_project_key(project: Optional[str]) -> str:
        normalized = (project or "").strip().lower()
        return normalized or "no project"

    @staticmethod
    def _display_project_name(project: Optional[str]) -> str:
        text = (project or "").strip()
        return text or "No Project"
    
    def generate_department_stats(
        self,
        all_user_data: Dict[str, Dict[str, Any]],
        users: List[User]
    ) -> Dict[str, Dict[str, Any]]:
        """Generate statistics by department."""
        
        department_stats = defaultdict(lambda: {
            'users': [],
            'total_hours': 0,
            'billable_hours': 0,
            'projects': set(),
            'absence_days': 0,
            'missing_days': 0
        })
        
        # Group users by department
        for user in users:
            department = user.department or 'Unknown'
            department_stats[department]['users'].append(user.email)
        
        # Aggregate data by department
        for user_email, user_data in all_user_data.items():
            user = next((u for u in users if u.email.lower() == user_email), None)
            department = user.department if user else 'Unknown'
            
            stats = department_stats[department]
            stats['total_hours'] += user_data.get('total_hours', 0)
            stats['billable_hours'] += user_data.get('billable_hours', 0)
            stats['absence_days'] += sum(user_data.get('absence_breakdown', {}).values())
            stats['missing_days'] += len(user_data.get('missing_days', []))
            
            # Add projects
            for project in user_data.get('project_hours', {}).keys():
                stats['projects'].add(project)
        
        # Convert to final format
        final_stats = {}
        for department, stats in department_stats.items():
            user_count = len(stats['users'])
            avg_hours_per_user = stats['total_hours'] / user_count if user_count > 0 else 0
            
            final_stats[department] = {
                'department_name': department,
                'user_count': user_count,
                'users': stats['users'],
                'total_hours': stats['total_hours'],
                'billable_hours': stats['billable_hours'],
                'average_hours_per_user': avg_hours_per_user,
                'total_projects': len(stats['projects']),
                'projects': list(stats['projects']),
                'total_absence_days': stats['absence_days'],
                'total_missing_days': stats['missing_days'],
                'billable_ratio': stats['billable_hours'] / stats['total_hours'] if stats['total_hours'] > 0 else 0,
                'productivity_score': self._calculate_department_productivity(stats)
            }
        
        return final_stats
    
    def _calculate_department_productivity(self, stats: Dict[str, Any]) -> float:
        """Calculate productivity score for a department."""
        if not stats['users']:
            return 0
        
        # Simple productivity calculation
        billable_ratio = stats['billable_hours'] / stats['total_hours'] if stats['total_hours'] > 0 else 0
        attendance_rate = 1 - (stats['missing_days'] / (len(stats['users']) * 30))  # Assuming 30 days per month
        
        return (billable_ratio * 0.6 + attendance_rate * 0.4) * 100
    
    def generate_summary_stats(
        self,
        all_user_data: Dict[str, Dict[str, Any]],
        users: List[User],
        year: int,
        month: int
    ) -> Dict[str, Any]:
        """Generate overall summary statistics."""
        
        total_users = len(users)
        active_users = len([data for data in all_user_data.values() if data.get('total_hours', 0) > 0])
        
        total_hours = sum(data.get('total_hours', 0) for data in all_user_data.values())
        total_billable = sum(data.get('billable_hours', 0) for data in all_user_data.values())
        total_absence_days = sum(sum(data.get('absence_breakdown', {}).values()) for data in all_user_data.values())
        total_missing_days = sum(len(data.get('missing_days', [])) for data in all_user_data.values())
        
        # Project statistics
        all_projects = set()
        for data in all_user_data.values():
            all_projects.update(data.get('project_hours', {}).keys())
        
        # Department statistics
        departments = set(user.department or 'Unknown' for user in users)
        
        # Calculate averages
        avg_hours_per_user = total_hours / total_users if total_users > 0 else 0
        avg_hours_per_active_user = total_hours / active_users if active_users > 0 else 0
        
        # Productivity metrics
        billable_ratio = total_billable / total_hours if total_hours > 0 else 0
        attendance_rate = 1 - (total_missing_days / (total_users * 30)) if total_users > 0 else 1
        
        return {
            'period': f"{year}-{month:02d}",
            'overview': {
                'total_users': total_users,
                'active_users': active_users,
                'total_departments': len(departments),
                'total_projects': len(all_projects),
                'departments': list(departments)
            },
            'hours_summary': {
                'total_hours': total_hours,
                'billable_hours': total_billable,
                'non_billable_hours': total_hours - total_billable,
                'billable_ratio': billable_ratio
            },
            'averages': {
                'avg_hours_per_user': avg_hours_per_user,
                'avg_hours_per_active_user': avg_hours_per_active_user,
                'avg_hours_per_day': total_hours / 30,  # Assuming 30 days per month
                'avg_billable_per_user': total_billable / total_users if total_users > 0 else 0
            },
            'attendance': {
                'total_absence_days': total_absence_days,
                'total_missing_days': total_missing_days,
                'attendance_rate': attendance_rate,
                'absence_rate': total_absence_days / (total_users * 30) if total_users > 0 else 0
            },
            'productivity': {
                'overall_productivity_score': (billable_ratio * 0.7 + attendance_rate * 0.3) * 100,
                'billable_ratio': billable_ratio,
                'attendance_rate': attendance_rate
            },
            'generated_at': datetime.now().isoformat()
        }
    
    def generate_trend_analysis(
        self,
        historical_data: List[Dict[str, Any]]
    ) -> Dict[str, Any]:
        """Generate trend analysis from historical data."""
        
        if len(historical_data) < 2:
            return {'error': 'Insufficient historical data for trend analysis'}
        
        # Sort by period
        sorted_data = sorted(historical_data, key=lambda x: (x['year'], x['month']))
        
        # Extract trends
        periods = [f"{d['year']}-{d['month']:02d}" for d in sorted_data]
        total_hours = [d['total_hours'] for d in sorted_data]
        billable_hours = [d['billable_hours'] for d in sorted_data]
        active_users = [d['active_users'] for d in sorted_data]
        
        # Calculate trends
        def calculate_trend(values):
            if len(values) < 2:
                return 0
            first_half = values[:len(values)//2]
            second_half = values[len(values)//2:]
            first_avg = sum(first_half) / len(first_half)
            second_avg = sum(second_half) / len(second_half)
            return ((second_avg - first_avg) / first_avg * 100) if first_avg > 0 else 0
        
        return {
            'periods': periods,
            'trends': {
                'total_hours_trend': calculate_trend(total_hours),
                'billable_hours_trend': calculate_trend(billable_hours),
                'active_users_trend': calculate_trend(active_users)
            },
            'latest_values': {
                'total_hours': total_hours[-1] if total_hours else 0,
                'billable_hours': billable_hours[-1] if billable_hours else 0,
                'active_users': active_users[-1] if active_users else 0
            },
            'data_points': len(sorted_data)
        }
    
    def export_statistics_summary(
        self,
        all_stats: Dict[str, Any]
    ) -> str:
        """Export statistics as a formatted summary string."""
        
        summary = []
        summary.append("=" * 60)
        summary.append("TIME TRACKING STATISTICS SUMMARY")
        summary.append("=" * 60)
        
        # Overview
        overview = all_stats.get('overview', {})
        summary.append(f"Period: {all_stats.get('period', 'Unknown')}")
        summary.append(f"Total Users: {overview.get('total_users', 0)}")
        summary.append(f"Active Users: {overview.get('active_users', 0)}")
        summary.append(f"Departments: {overview.get('total_departments', 0)}")
        summary.append(f"Projects: {overview.get('total_projects', 0)}")
        summary.append("")
        
        # Hours summary
        hours = all_stats.get('hours_summary', {})
        summary.append("HOURS SUMMARY:")
        summary.append(f"  Total Hours: {hours.get('total_hours', 0):.1f}")
        summary.append(f"  Billable Hours: {hours.get('billable_hours', 0):.1f}")
        summary.append(f"  Non-billable Hours: {hours.get('non_billable_hours', 0):.1f}")
        summary.append(f"  Billable Ratio: {hours.get('billable_ratio', 0):.1%}")
        summary.append("")
        
        # Averages
        averages = all_stats.get('averages', {})
        summary.append("AVERAGES:")
        summary.append(f"  Hours per User: {averages.get('avg_hours_per_user', 0):.1f}")
        summary.append(f"  Hours per Active User: {averages.get('avg_hours_per_active_user', 0):.1f}")
        summary.append(f"  Hours per Day: {averages.get('avg_hours_per_day', 0):.1f}")
        summary.append("")
        
        # Productivity
        productivity = all_stats.get('productivity', {})
        summary.append("PRODUCTIVITY:")
        summary.append(f"  Overall Score: {productivity.get('overall_productivity_score', 0):.1f}")
        summary.append(f"  Attendance Rate: {productivity.get('attendance_rate', 0):.1%}")
        summary.append("")
        
        summary.append("=" * 60)
        
        return "\n".join(summary)
